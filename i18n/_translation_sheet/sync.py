"""
translations.xlsx を i18n/ja.js と i18n/en.js から再生成する。

使い方（プロジェクトルートで実行）:
    python i18n/_translation_sheet/sync.py

方針（マージ）:
- ja / en とも、コード側（*.js）に値があればそれを最優先で書き出す。
- コード側が null で、既存 xlsx に値があればそれを残す（下書き保護）。
- どちらにも無ければ空セルにする。
- キーの並びは ja.js の走査順。ja に無く en にだけあるキーは末尾に置く。

関数エントリ（テンプレート文字列）は、プロパティ名がそのまま
`${name}` として埋め込まれた形で書き出される（例: `${key}`, `${n}`）。
"""

import json
import subprocess
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[2]
JA_JS = ROOT / "i18n" / "ja.js"
EN_JS = ROOT / "i18n" / "en.js"
OUT_XLSX = ROOT / "i18n" / "_translation_sheet" / "translations.xlsx"

NODE_SCRIPT = r"""
const path = require('path');
const window = {};
global.window = window;
require(path.resolve(process.argv[1]));
require(path.resolve(process.argv[2]));
const dicts = window.KIOKU_PIN_I18N || {};

function flatten(obj, prefix) {
  const out = {};
  if (obj == null || typeof obj !== 'object') return out;
  for (const k of Object.keys(obj)) {
    const v = obj[k];
    const key = prefix ? prefix + '.' + k : k;
    if (v === null || v === undefined) {
      out[key] = null;
    } else if (typeof v === 'function') {
      const proxy = new Proxy({}, { get: (_, p) => '${' + String(p) + '}' });
      try { out[key] = v(proxy); } catch (e) { out[key] = null; }
    } else if (typeof v === 'object') {
      Object.assign(out, flatten(v, key));
    } else {
      out[key] = String(v);
    }
  }
  return out;
}

const ja = flatten(dicts.ja || {}, '');
const en = flatten(dicts.en || {}, '');
const order = Object.keys(ja);
for (const k of Object.keys(en)) if (!(k in ja)) order.push(k);
process.stdout.write(JSON.stringify({ order, ja, en }));
"""


def extract_from_js() -> dict:
    proc = subprocess.run(
        ["node", "-e", NODE_SCRIPT, str(JA_JS), str(EN_JS)],
        capture_output=True,
    )
    if proc.returncode != 0:
        sys.stderr.write(proc.stderr.decode("utf-8", errors="replace"))
        raise SystemExit(proc.returncode)
    return json.loads(proc.stdout.decode("utf-8"))


def load_existing_xlsx() -> dict:
    if not OUT_XLSX.exists():
        return {}
    wb = openpyxl.load_workbook(OUT_XLSX)
    ws = wb.active
    existing = {}
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return {}
    for row in rows:
        if not row or not row[0]:
            continue
        key = row[0]
        ja = row[1] if len(row) > 1 else None
        en = row[2] if len(row) > 2 else None
        existing[key] = {"ja": ja, "en": en}
    return existing


def merge(code: dict, existing: dict) -> list:
    """コード側優先、code が null なら xlsx の既存値を残す。"""
    rows = []
    for key in code["order"]:
        prev = existing.get(key, {})
        ja = code["ja"].get(key)
        if ja is None:
            ja = prev.get("ja")
        en = code["en"].get(key)
        if en is None:
            en = prev.get("en")
        rows.append((key, ja, en))
    return rows


def write_xlsx(rows: list) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "translations"
    ws.append(["key", "ja", "en"])
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="EEEEEE")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in rows:
        ws.append(list(row))
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 60
    ws.freeze_panes = "A2"
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)


def main() -> None:
    code = extract_from_js()
    existing = load_existing_xlsx()
    rows = merge(code, existing)
    write_xlsx(rows)
    filled_en = sum(1 for _, _, en in rows if en)
    print(f"wrote {OUT_XLSX} ({len(rows)} keys, en filled: {filled_en})")


if __name__ == "__main__":
    main()
